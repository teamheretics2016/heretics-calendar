#!/usr/bin/env python3
"""
Team Heretics — Parser de Calendario Deportivo
Convierte Calendario_deportivo_2026.xlsx → events.json

Uso:
  python parse_excel.py <ruta_al_excel> <ruta_salida_json>

Ejemplo:
  python parse_excel.py Calendario_deportivo_2026.xlsx ../web/events.json
"""

import sys
import json
import re
import math
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: Necesitas instalar openpyxl: pip install openpyxl")
    sys.exit(1)

# ─────────────────────────────────────────────────────────────────────
# CONFIGURACIÓN
# ─────────────────────────────────────────────────────────────────────

SPORT_COLORS = {
    "FFBF9000": "LEC",
    "FF00FFFF": "CDL",
    "FF674EA7": "SL",
    "FFEA9999": "VCT",
    "FFFF00FF": "BRAWL",
    "FF999999": "R6",
    "FF3C78D8": "MARVEL",
}

MONTH_NAMES = {
    "ENERO": 1, "FEBRERO": 2, "MARZO": 3, "ABRIL": 4,
    "MAYO": 5, "JUNIO": 6, "JULIO": 7, "AGOSTO": 8,
    "SEPTIEMBRE": 9, "OCTUBRE": 10, "NOVIEMBRE": 11, "DICIEMBRE": 12,
}

DAY_PREFIXES = ["L", "M", "X", "J", "V", "S", "D",
                "LU", "MA", "MI", "JU", "VI", "SA", "DO"]

# Partidos añadidos manualmente (no están en el Excel)
MANUAL_EVENTS = [
    {"year": 2026, "month": "FEBRERO", "month_num": 2, "day": 13,
     "sport": "SL", "competition": "Kickoff LES", "time": "19:00",
     "opponent": "Barça Esports"},
    {"year": 2026, "month": "FEBRERO", "month_num": 2, "day": 15,
     "sport": "SL", "competition": "Kickoff LES", "time": "15:00",
     "opponent": "KOI Fénix"},
    {"year": 2026, "month": "FEBRERO", "month_num": 2, "day": 21,
     "sport": "LEC", "competition": "LEC Winter Playoffs W9", "time": "16:45",
     "opponent": "GIANTX"},
]

# ─────────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────────

def get_color(cell):
    try:
        fill = cell.fill
        if fill and fill.fgColor:
            rgb = fill.fgColor.rgb
            if isinstance(rgb, str) and len(rgb) == 8 and rgb != "00000000":
                return rgb.upper()
    except Exception:
        pass
    return None


def get_comment(cell):
    try:
        if cell.comment:
            return cell.comment.text.strip()
    except Exception:
        pass
    return None


def parse_comment(text):
    """Extrae (time, opponent) del texto de comentario."""
    if not text:
        return None, None
    # Eliminar timezone: " | 4:30pm ET", " (3pm ET)", etc.
    text = re.sub(r'\s*[\|]\s*\d+:?\d*\s*(?:am|pm)\s*\w*', '', text, flags=re.IGNORECASE)
    text = re.sub(r'\s*\(\d+:?\d*\s*(?:am|pm)\s*\w*\)', '', text, flags=re.IGNORECASE)
    text = text.strip()
    # Formato: "HH:MM vs Rival" (con o sin espacio después de vs)
    m = re.match(r'^(\d{1,2}:\d{2})\s*vs\s*(.+)', text, re.IGNORECASE)
    if m:
        time_str = m.group(1)
        opponent = m.group(2).strip()
        # Limpiar nombre rival
        opponent = re.sub(r'\s*\((?:BO\d+|\d+(?::\d+)?(?:am|pm)\s*\w*)\)', '', opponent, flags=re.IGNORECASE).strip()
        # TBD → None
        if opponent.upper() in ("TBD", "TBA", "???", ""):
            opponent = None
        # Rebranding: LA Guerrillas M8 → Paris M8
        if opponent and "guerrillas m8" in opponent.lower():
            opponent = "Paris M8"
        return time_str, opponent
    return None, None


def is_day_header(cell_value):
    """¿Es una cabecera de día? Ej: 'L 05', 'MA 12'"""
    if not isinstance(cell_value, str):
        return False
    parts = cell_value.strip().split()
    if len(parts) == 2:
        prefix = parts[0].upper()
        num = parts[1]
        if prefix in DAY_PREFIXES and num.isdigit():
            return True
    return False


def extract_day_num(cell_value):
    """Extrae el número de día de 'L 05' → 5"""
    parts = str(cell_value).strip().split()
    if len(parts) >= 2:
        try:
            return int(parts[1])
        except ValueError:
            pass
    return None


def fix_cdl_midnight(events):
    """CDL con hora 00:xx ó 01:xx → avanzar al día siguiente (ya es España)."""
    import calendar
    midnight_hours = {"00:00", "00:30", "01:00", "01:30"}
    fixed = 0
    for e in events:
        if e["sport"] == "CDL" and e.get("time") in midnight_hours:
            day = e["day"]
            month = e["month_num"]
            year = e["year"]
            # Avanzar día
            max_day = calendar.monthrange(year, month)[1]
            if day < max_day:
                e["day"] = day + 1
            else:
                e["day"] = 1
                if month < 12:
                    e["month_num"] = month + 1
                    e["month"] = [k for k, v in MONTH_NAMES.items() if v == month + 1][0]
                else:
                    e["month_num"] = 1
                    e["month"] = "ENERO"
                    e["year"] = year + 1
            fixed += 1
    print(f"  CDL midnight fix: {fixed} eventos")
    return events


def fix_brawl_time(events):
    """Brawl sin hora → 14:00"""
    for e in events:
        if e["sport"] == "BRAWL" and not e.get("time"):
            e["time"] = "14:00"
    return events


def compute_format(e):
    """Calcula Bo1/Bo3/Bo5 según sport y fase."""
    sport = e["sport"]
    comp = (e.get("competition") or "").lower()

    if sport == "CDL":
        return "Bo5"
    if sport == "LEC":
        return "Bo5" if any(x in comp for x in ["playoff", "po", "final"]) else "Bo1"
    if sport == "VCT":
        if any(x in comp for x in ["grand final", "gran final", "gf"]):
            return "Bo5"
        if any(x in comp for x in ["lower final", "upper final", "ub final", "lb final", "playoff final"]):
            return "Bo5"
        if any(x in comp for x in ["lower", "upper", "bracket", "playoff"]):
            return "Bo3"
        return "Bo1"
    if sport == "SL":
        if any(x in comp for x in ["gran final", "grand final", "gf"]):
            return "Bo5"
        if any(x in comp for x in ["playoff", "po", "semifinal", "cuartos"]):
            return "Bo3"
        return "Bo1"
    if sport == "R6":
        return "Bo3" if any(x in comp for x in ["playoff", "final"]) else "Bo1"
    if sport == "MARVEL":
        return "Bo3"
    if sport == "BRAWL":
        return "Bo3"
    return "Bo1"


def compute_split_key(e):
    """Genera la clave que agrupa una fase/torneo."""
    sport = e["sport"]
    year = e["year"]
    comp = (e.get("competition") or "").lower()

    if sport == "CDL":
        for i, kw in enumerate(["qualifier 1", "q1", "qualifier 2", "q2",
                                  "qualifier 3", "q3", "qualifier 4", "q4"]):
            if kw in comp:
                q = ["Q1", "Q2", "Q3", "Q4"][i // 2]
                return f"CDL_{q}_{year}"
        for kw, tag in [("major i", "MAJOR_I"), ("major ii", "MAJOR_II"),
                         ("major iii", "MAJOR_III"), ("minor", "MINOR"),
                         ("champ", "CHAMPS")]:
            if kw in comp:
                return f"CDL_{tag}_{year}"
        return f"CDL_OTHER_{year}"

    if sport == "LEC":
        for kw, tag in [("winter", "WINTER"), ("spring", "SPRING"), ("summer", "SUMMER")]:
            if kw in comp:
                return f"LEC_{tag}_{year}"
        return f"LEC_OTHER_{year}"

    if sport == "VCT":
        for kw, tag in [("kickoff", "KICKOFF"), ("stage 1", "STAGE1"), ("stage 2", "STAGE2"),
                         ("masters", "MASTERS"), ("champion", "CHAMPS")]:
            if kw in comp:
                return f"VCT_{tag}_{year}"
        return f"VCT_OTHER_{year}"

    if sport == "SL":
        if "emea" in comp:
            for kw, tag in [("winter", "EMEA_WINTER"), ("spring", "EMEA_SPRING"), ("summer", "EMEA_SUMMER")]:
                if kw in comp:
                    return f"SL_{tag}_{year}"
            return f"SL_EMEA_{year}"
        for kw, tag in [("kickoff", "KICKOFF"), ("winter", "WINTER"),
                         ("spring", "SPRING"), ("summer", "SUMMER")]:
            if kw in comp:
                return f"SL_{tag}_{year}"
        return f"SL_OTHER_{year}"

    if sport == "BRAWL":
        return f"BRAWL_{year}"
    if sport == "R6":
        return f"R6_{year}"
    if sport == "MARVEL":
        if "pre" in comp or "preseason" in comp:
            return f"MARVEL_PRESEASON_{year}"
        for kw, tag in [("stage 1", "STAGE1"), ("stage 2", "STAGE2")]:
            if kw in comp:
                return f"MARVEL_{tag}_{year}"
        return f"MARVEL_OTHER_{year}"

    return f"{sport}_OTHER_{year}"


def add_jornadas(events):
    """Añade _jornada y _split_key. Ordena primero."""
    events.sort(key=lambda e: (e["year"], e["month_num"], e["day"], e.get("time") or "99:99"))
    counters = {}
    for e in events:
        sk = compute_split_key(e)
        e["_split_key"] = sk
        counters[sk] = counters.get(sk, 0) + 1
        e["_jornada"] = counters[sk]
    return events


# ─────────────────────────────────────────────────────────────────────
# PARSER PRINCIPAL
# ─────────────────────────────────────────────────────────────────────

def parse_sheet(ws, year):
    """Parsea una hoja del Excel y devuelve lista de eventos."""
    events = []
    # Ajuste especial para 2022 (columnas desplazadas)
    if year == 2022:
        left_col_start, right_col_start = 4, 12
    else:
        left_col_start, right_col_start = 2, 10

    max_row = ws.max_row
    max_col = ws.max_column

    # 1. Encontrar bloques de meses
    month_blocks = []
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.strip().upper() in MONTH_NAMES:
                mn = cell.value.strip().upper()
                mc = cell.column
                mr = cell.row
                # Determinar si es mes izquierdo o derecho
                if mc <= (left_col_start + 4):
                    col_start = left_col_start
                else:
                    col_start = right_col_start
                col_end = col_start + 6  # CRÍTICO: exactamente 7 columnas
                month_blocks.append({
                    "name": mn,
                    "num": MONTH_NAMES[mn],
                    "row_start": mr,
                    "col_start": col_start,
                    "col_end": col_end,
                })

    # Determinar row_end de cada bloque
    for i, mb in enumerate(month_blocks):
        if i + 1 < len(month_blocks):
            same_col = [b for b in month_blocks[i+1:] if b["col_start"] == mb["col_start"]]
            if same_col:
                mb["row_end"] = same_col[0]["row_start"] - 1
            else:
                mb["row_end"] = max_row
        else:
            mb["row_end"] = max_row

    # 2. Para cada celda con color y comentario, encontrar mes y día
    all_cells = []
    for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
        for cell in row:
            color = get_color(cell)
            comment = get_comment(cell)
            if color in SPORT_COLORS and comment:
                all_cells.append((cell, color, comment))

    for cell, color, comment in all_cells:
        cr, cc = cell.row, cell.column
        sport = SPORT_COLORS[color]
        time_str, opponent = parse_comment(comment)

        # Encontrar mes por containment
        month_info = None
        for mb in month_blocks:
            if (mb["row_start"] <= cr <= mb["row_end"] and
                    mb["col_start"] <= cc <= mb["col_end"]):
                month_info = mb
                break
        if not month_info:
            continue

        # Encontrar día: buscar cabecera de día encima en la misma columna
        day_num = None
        competition = None
        for check_row in range(cr - 1, month_info["row_start"], -1):
            check_cell = ws.cell(row=check_row, column=cc)
            val = check_cell.value
            if val and is_day_header(str(val)):
                day_num = extract_day_num(str(val))
                break
            # La celda podría tener el nombre de la competición
            if val and isinstance(val, str) and not competition:
                competition = val.strip()

        # Alternativamente buscar la competición en la propia celda
        if not competition and cell.value:
            competition = str(cell.value).strip()

        if day_num is None:
            continue

        events.append({
            "year": year,
            "month": month_info["name"],
            "month_num": month_info["num"],
            "day": day_num,
            "sport": sport,
            "competition": competition or "",
            "time": time_str,
            "opponent": opponent,
        })

    return events


def parse_excel(xlsx_path):
    print(f"Abriendo: {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    print(f"Hojas encontradas: {wb.sheetnames}")

    all_events = []

    for sheet_name in wb.sheetnames:
        try:
            year = int(sheet_name)
        except ValueError:
            print(f"  Saltando hoja '{sheet_name}' (no es un año)")
            continue

        ws = wb[sheet_name]
        print(f"  Parseando {year}...")
        sheet_events = parse_sheet(ws, year)
        print(f"    → {len(sheet_events)} eventos encontrados")
        all_events.extend(sheet_events)

    # Añadir manuales
    for e in MANUAL_EVENTS:
        all_events.append(dict(e))
    print(f"  Manuales añadidos: {len(MANUAL_EVENTS)}")

    # Correcciones
    all_events = fix_cdl_midnight(all_events)
    all_events = fix_brawl_time(all_events)

    # Añadir jornadas y split keys
    all_events = add_jornadas(all_events)

    # Añadir formato Bo
    for e in all_events:
        e["format"] = compute_format(e)

    print(f"\nTotal eventos: {len(all_events)}")
    return all_events


# ─────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    if len(sys.argv) < 3:
        print(__doc__)
        sys.exit(1)

    xlsx_path = sys.argv[1]
    out_path = sys.argv[2]

    if not Path(xlsx_path).exists():
        print(f"ERROR: No se encuentra el archivo: {xlsx_path}")
        sys.exit(1)

    events = parse_excel(xlsx_path)

    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(events, f, ensure_ascii=False, separators=(",", ":"))

    print(f"✓ Guardado en: {out_path}")
    print(f"  {len(events)} eventos en {len(set(e['year'] for e in events))} años")
